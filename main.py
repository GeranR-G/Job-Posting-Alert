from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from classes import Posting
from mapping import EMPLOYER, JOB, URL, POSTING_DATE, EXPIRY_DATE
from config import email_list, alert_days, snapshot_days, file_path, date_format

def main():
    today = date.today()
    jobs = get_jobs()
    
def get_jobs(): #creates a list of job posting objects
    jobs = []
    sheet = load_workbook(filename=file_path, read_only=True).active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            job = Posting(employer=row[EMPLOYER],
                        job= row[JOB],
                        url= row[URL],
                        date_posted= datetime.strptime(row[POSTING_DATE], date_format).date(),
                        expiry_date= datetime.strptime(row[EXPIRY_DATE], date_format).date())
            jobs.append(job)
    return jobs

def adjust_for_weekends(date: date): #adjusts the date if it lands on a weekend
    if date.weekday() == 5:
        return date - timedelta(days=1)
    elif date.weekday() == 6:
        return date - timedelta(days=2)
    return date

main()
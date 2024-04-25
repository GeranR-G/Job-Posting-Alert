import smtplib, ssl, config, mapping, sys
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from classes import Posting
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from config import ConfigSetting

CONFIGSETTINGS = ConfigSetting()

def main():
    today = date.today()
    jobs = get_jobs()
    message = compose_message(jobs, today)
    if len(message) != 0:
        send_message(message)
    
def get_jobs(): #creates a list of job posting objects
    jobs = []
    print(CONFIGSETTINGS.file_path)
    sheet = load_workbook(filename=CONFIGSETTINGS.file_path, read_only=True).active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            job = Posting(employer=row[mapping.EMPLOYER],
                        job= row[mapping.JOB],
                        url= row[mapping.URL],
                        date_posted= datetime.strptime(row[mapping.POSTING_DATE], "%Y-%m-%d").date(),
                        expiry_date= datetime.strptime(row[mapping.EXPIRY_DATE], "%Y-%m-%d").date())
            jobs.append(job)
    return jobs

def adjust_for_weekends(date): #adjusts the date if it lands on a weekend
    if date.weekday() == 5:
        return date - timedelta(days=1)
    elif date.weekday() == 6:
        return date - timedelta(days=2)
    return date

def add_days(date, modifier): #adds days to a given date
    return date + timedelta(days=modifier)

def check_for_snapshot_bool(job, today): #checks if a snapshot needs to be taken
    snapshot_date = adjust_for_weekends(add_days(job.date_posted, CONFIGSETTINGS.snapshot_days))
    return snapshot_date >= today and snapshot_date <= add_days(today, CONFIGSETTINGS.alert_days)

def check_for_expiry_bool(job, today): #checks if the posting is about to expire
    expiry_date = adjust_for_weekends(job.expiry_date)
    return expiry_date >= today and expiry_date <= add_days(today, CONFIGSETTINGS.alert_days)

def add_actions(job, today):
    message = ""
    if check_for_snapshot_bool(job, today):
        message += f"\n- The {job.job} posting needs a {CONFIGSETTINGS.snapshot_days} day snapshot taken on {adjust_for_weekends(add_days(job.date_posted, config.snapshot_days))},  url: {job.url}"
    if check_for_expiry_bool(job, today):
        message += f"\n- The {job.job} posting will expire on {job.expiry_date} and the last day to renew it is {adjust_for_weekends(job.expiry_date)}, url: {job.url}"
    return message

    
def compose_message(jobs, today):
    old_employer = ""
    message = ""
    for job in jobs:
        action_needed = check_for_expiry_bool(job, today) or check_for_snapshot_bool(job, today)
        if message == "" and action_needed:
            message = f"Job(s) for {job.employer} need the following update(s): "
            old_employer = job.employer
            message += add_actions(job, today)
        elif old_employer == job.employer and action_needed:
            message += add_actions(job, today)
        elif old_employer != job.employer and action_needed:
            message += f"\nJob(s) for {job.employer} need the following update(s): "
            old_employer = job.employer
            message += add_actions(job, today)
    return message

def send_message(message_text):
    port = 465
    smpt_server = "smtp.gmail.com"
    password = input("Type your password and press enter: ")
    message = MIMEMultipart("alternative")
    message["Subject"] = "Job post(s) need attention"
    message["From"] = CONFIGSETTINGS.your_gmail
    text = MIMEText(message_text, "plain")
    message.attach(text)
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(smpt_server, port, context=context) as server:
        server.login(CONFIGSETTINGS.your_gmail, password)
        server.sendmail(CONFIGSETTINGS.your_gmail, CONFIGSETTINGS.your_gmail, message.as_string())
        for email in CONFIGSETTINGS.email_list:
            server.sendmail(CONFIGSETTINGS.your_gmail, email, message.as_string())

main()